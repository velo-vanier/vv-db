import openpyxl
import pymysql
import logging

logger = logging.getLogger()
logger.setLevel(logging.INFO)

host  = 'bikesharemysql.cwf839dt8hxr.us-east-2.rds.amazonaws.com'
name = 'BikeShareMaster'
password = 'MasterBikeShare2018'
db_name = 'BikeShare'

logger.info(host, name, password, db_name)

def addPhoto(bikeLabel, photoLink):
    conn = pymysql.connect(host, user=name, passwd=password, db=db_name, connect_timeout=5)

    with conn.cursor() as cur:
        
        # make sure this bike exists, it doesn't do ignore
        sql = 'select ID_Bike from Bike where BikeLabel = "{}"'.format(bikeLabel)
        logger.info(sql)

        cur.execute(sql)
        row = cur.fetchone()
        if row:
            idBike = row[0]
            logger.info("bike exists - bikeLabel: {}".format(idBike))

            # check if photo exists
            cur.execute("select ID_Photo from Photo where url = '{}'".format(photoLink))
            logger.info(sql)
            row = cur.fetchone()
            if not row:
                print("insert new photo")
                cur.excute("INSERT INTO `BikeShare`.`Photo` (`url`) VALUES '{}');".format(photoLink))
                conn.commit()
                idPhoto = conn.insert_id()
            else:
                idPhoto = row[0]

            # TODO
            # check that bike/photo doesn't exist
            # unique constraint is there for now

            logger.info("insert new bike/photo (ID_Bike: {}, ID_Photo: {})".format(idBike, idPhoto))
            sql = "INSERT INTO `BikeShare`.`BikePhoto` (ID_Bike, ID_Photo) VALUES ({}, {});".format(idBike, idPhoto)
            cur.excute(sql)
            conn.commit()

wb = openpyxl.load_workbook('/home/rob/rhok/velo-vanier/Velo-Vanier Operations Log.xlsx')
ws = wb.get_sheet_by_name('CalculateStatus')

print(ws.cell(row=3, column=1).hyperlink.target)
maxRow=ws.max_row
maxColumn=ws.max_column

print(maxRow,maxColumn)

for i in range(1, maxRow + 1):
    try:
        cell = ws.cell(row=i, column=1)

        bikeLabel = cell.value
        bikeLink = cell.hyperlink.target    

        print(bikeLabel)
        print(bikeLink)     

        addPhoto(bikeLabel, bikeLink)

    except Exception as e:
        print(e)
